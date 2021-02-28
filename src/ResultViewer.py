import os
import wx
import wx.adv
from wx.lib.itemspicker import ItemsPicker,EVT_IP_SELECTION_CHANGED, IP_SORT_CHOICES
from wx.lib.itemspicker import IP_SORT_SELECTED,IP_REMOVE_FROM_CHOICES
import traceback
import configparser as ConfigParser  # keep compytibility with python 2.7
import glob
import re
import shutil
import datetime
from Plugins import create_html
from threading import Thread
import zipfile
import socket
import wx.lib.agw.ribbon as RB
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import getpass
import collections
import logging
from Plugins import ReportHTML
import scandir


# Create the log file for the application
log_file_name = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + ".log"
# logging.basicConfig(filename='.data/Log/' + log_file_name, filemode='w', level=logging.DEBUG, format='%(asctime)s\t%(levelname)s\t\t%(message)s')
# logging.debug("====================")
# logging.debug("Starting application")


class Setting:
    def __init__(self, path):
        self.path = path
        reset_needed = False

        if not os.path.isfile(self.path):
            setting_f = open(self.path, "w")
            setting_f.close()
            reset_needed = True

        self.Config = ConfigParser.ConfigParser()
        self.Config.read(self.path)
        if reset_needed:
            self.reset()

    def reset(self):
        self.Config.add_section("file")
        self.Config.set(section="file", option="read folder", value="")
        self.Config.set(section="file", option="number of readed files", value="200")
        self.Config.set(section="file", option="store folder", value="")
        self.Config.add_section("windows")
        self.Config.set(section="windows", option="metric panel", value="True")
        self.Config.set(section="windows", option="messages panel", value="True")

        setting_f = open(self.path, "w")
        self.Config.write(setting_f)
        setting_f.close()

    def read(self, section="", option=""):
        return self.Config.get(section=section, option=option)

    def write(self, section="", option="", value=""):
        self.Config.read(self.path)
        self.Config.set(section=section, option=option, value=value)

        file_to_write = open(self.path, "w")
        self.Config.write(file_to_write)
        file_to_write.close()


class TestResultParser:
    def __init__(self, path, limit_number_of_file = -1, recursive=False):
        self.dizionario = {}
        self.errors = []
        PATH = path

        # read all file in folder and create a list
        list_of_f = []
        if not recursive:
            for f in glob.glob(PATH + r"\*.txt"):
                list_of_f.append(f)
        elif recursive:
            for root, dirs, files in scandir.walk(PATH):
                for f in files:
                    if f.endswith(".txt"):
                        list_of_f.append(root + "\\" + f)

        # sort the list from the newer to older
        list_of_f.sort(reverse=True)

        # keep the first "n" files
        if limit_number_of_file != -1:
            list_of_f = list_of_f[:limit_number_of_file]

        thrds = []
        i = 1
        for f in list_of_f:
            t = Thread(target=self.reader, args=(f, i))
            thrds.append(t)
            i += 1
            # each 50 files, execute the threads
            if i%50 == 0:
                # start the threads
                for t in thrds:
                    t.start()
                # wait until all threads are finished
                for t in thrds:
                    t.join()
                # clean up the list of threads
                del thrds[:]

        # check if there are remaining thread to start
        if thrds:
            # start the threads
            for t in thrds:
                t.start()
            # wait until all threads are finished
            for t in thrds:
                t.join()
            # clean up the list of threads
            del thrds[:]

    def reader(self, f, i):
        try:
            parser = ConfigParser.ConfigParser()
            parser.read(f)
            self.dizionario[i] = parser._sections
            self.dizionario[i]["GENERIC"]["path"] = f
        except:
            self.errors.append(f)
            print("ERROR: error raised with TC: " + f)

    def get_dictionary(self):
        return self.dizionario, self.errors


class Report:
    @staticmethod
    def Overview(reportDirectory, files):
        """

        :param reportDirectory:
        :param files:
        :return: 0= no problem, 1=missing info, 2=manual test detected, 3=missing info + manual test detected
        """
        print("=== Overview report started ===")
        # check the directory exist
        if not os.path.isdir(reportDirectory):
            print("Folder choosen does not exist")
            return -1

        # Initialization
        TestReportTitle = None
        ReportFileName = None
        ReportWorkbook = None
        ReportWorksheet = None
        CaseNumber = 0
        RowIndex = 3

        Config = ConfigParser.ConfigParser()

        # Store the title
        TestReportTitle = "Result"
        # Create the prefix of the report file name
        today = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
        ReportFileName = reportDirectory + r"\\" + today + "_Overview.xlsx"

        # Initialization of the report file in Excel
        # Open the workbook
        ReportWorkbook = Workbook()
        ReportWorksheet = ReportWorkbook.worksheets[0]

        ReportWorksheet.title = TestReportTitle

        # set the header and footer of the page and page orientation
        # ReportWorksheet.header_footer.right_header.text = "&D"     # Add the date in the right of the header
        # ReportWorksheet.header_footer.left_header.text = "&A"      # Add name of worksheet
        ReportWorksheet.header_footer.right_footer.text = "LIEBHERR MACHINE BULLE"
        ReportWorksheet.header_footer.left_footer.text = "&D\nDoc ID: &[File]"
        ReportWorksheet.header_footer.center_footer.text = "&[Page]/&[Pages]"
        ReportWorksheet.page_setup.orientation = ReportWorksheet.ORIENTATION_LANDSCAPE

        # Set the first three row to be reapeated at the top on each printed page
        ReportWorksheet.add_print_title(3)

        # Set the Main title
        localCell = ReportWorksheet.cell('A1')
        localCell.value = "Result Overview"
        # Font properties
        localCell.font = Font(name='Calibri', size=16, bold=True)
        localCell.alignment = Alignment(horizontal="center")
        # Cell background
        localCell.fill = PatternFill(fill_type="solid", start_color="00F5F6F7")
        # Merge the cells
        ReportWorksheet.merge_cells("A1:I1")

        "No more necessary because initialise during the merge above"
        # Initialise Column F
        # localCell = ReportWorksheet.cell('F1')
        # localCell.value = ""

        # Initialise Column G
        localCell = ReportWorksheet.cell('G1')
        localCell.value = ""

        # Initialise Column H
        localCell = ReportWorksheet.cell('H1')
        localCell.value = ""

        # Initialise Column I
        localCell = ReportWorksheet.cell('I1')
        localCell.value = ""

        # Set the column width
        ReportWorksheet.column_dimensions['A'].width = 50
        ReportWorksheet.column_dimensions['B'].width = 50
        ReportWorksheet.column_dimensions['C'].width = 40
        ReportWorksheet.column_dimensions['D'].width = 15
        ReportWorksheet.column_dimensions['E'].width = 15
        ReportWorksheet.column_dimensions['F'].width = 20
        ReportWorksheet.column_dimensions['G'].width = 15
        ReportWorksheet.column_dimensions['H'].width = 70
        ReportWorksheet.column_dimensions['I'].width = 20
        ReportWorksheet.column_dimensions['J'].width = 40

        list_of_title = [("TestCase ID", "A3"),
                         ("Requirement ID ", "B3"),
                         ("Description", "C3"),
                         ("SW No.", "D3"),
                         ("Revision No.", "E3"),
                         ("Date Tested", "F3"),
                         ("Result", "G3"),
                         ("Comment", "H3"),
                         ("Incident No.", "I3"),
                         ("Add. info.", "J3")]

        for t in list_of_title:
            localCell = ReportWorksheet.cell(t[1])
            # insert bottom border
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
            localCell.value = t[0]
            # Font properties
            localCell.font = Font(name='Calibri', size=13, bold=True)
            localCell.alignment = Alignment(horizontal="left")

        ERROR_STATUS_TC_FOUND = False
        ERROR_MISSING_INFO_IN_TC = False
        for report in files:
            Config.read(report)
            GREY_COLOR = "00ECECEC"

            RowIndex += 1
            Column = 1

            values = [("GENERIC", "test case number"),
                      ("GENERIC", "requirement cover"),
                      ("GENERIC", "test case description"),
                      ("ENVIRONMENT", "sw number"),
                      ("ENVIRONMENT", "sw revision"),
                      ("GENERIC", "test execution date"),
                      ("GENERIC", "result"),
                      ("GENERIC", "comment"),
                      ("GENERIC", "incident number"),
                      ("ENVIRONMENT", "can1 protocol"),
                      ("plus 1", "plus 1"),
                      ("svn link", "svn link")]
            for tc in values:
                localCell = ReportWorksheet.cell(row=RowIndex, column=Column)

                # color on line each two
                # if RowIndex % 2 == 0:
                #     localCell.fill = PatternFill(fill_type="solid", start_color=GREY_COLOR)

                # insert bottom border
                localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))

                # set the vertical alignment to center
                localCell.alignment = Alignment(vertical="center")
                try:
                    if tc[1] == "plus 1":
                        value_to_write = 1
                    elif tc[1] == "svn link":
                        value_to_write = "ver" # SVN.GetURL(report)
                    else:
                        value_to_write = Config.get(section=tc[0], option=tc[1])

                    if tc[1] == "test case number":
                        localCell.alignment = Alignment(wrap_text=True, vertical="center")
                    elif tc[1] == "requirement cover":
                        localCell.alignment = Alignment(wrap_text=True, vertical="center")
                        value_to_write = value_to_write.lstrip()
                    elif tc[1] == "result":
                        if value_to_write == "NOT OK":
                            localCell.fill = PatternFill(fill_type="solid", start_color="00FF0000")
                            if Config.get(section="GENERIC", option="comment") == "" or Config.get(section="GENERIC", option="incident number") == "":
                                ERROR_MISSING_INFO_IN_TC = True
                        elif value_to_write == "OK":
                            localCell.fill = PatternFill(fill_type="solid", start_color="0000FF00")
                        elif value_to_write == "NOT TESTED":
                            if Config.get(section="GENERIC", option="comment") == "":
                                ERROR_MISSING_INFO_IN_TC = True
                        else:
                            ERROR_STATUS_TC_FOUND = True

                        localCell.alignment = Alignment(horizontal="center", vertical="center")
                    elif tc[1] == "test case description":
                        localCell.alignment = Alignment(wrap_text=True, vertical="center")
                    elif tc[1] == "test execution date":
                        time_executed = Config.get(section="GENERIC", option="test execution time")
                        value_to_write = value_to_write + "\n " + time_executed
                        localCell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    elif tc[1] == "sw number" or tc[1] == "sw revision":
                        localCell.alignment = Alignment(horizontal="center", vertical="center")
                    elif tc[1] == "can1 protocol":
                        can2_protocol = Config.get(section="ENVIRONMENT", option="can2 protocol")
                        engine_variant = Config.get(section="ENVIRONMENT", option="engine variant")
                        value_to_write = "CAN1 prot: " + value_to_write + "\n" + "CAN2 prot: " + can2_protocol + "\n" + "Eng. vari.: " + engine_variant
                        localCell.alignment = Alignment(wrap_text=True)

                    localCell.value = value_to_write
                except:
                    localCell.value = "NA"
                    print("______________________________")
                    print("Error with TC: " + report)
                    print(traceback.format_exc(0))

                Column += 1

        # Save the report
        ReportWorkbook.save(ReportFileName)

        # Return value
        if ERROR_MISSING_INFO_IN_TC and ERROR_STATUS_TC_FOUND:
            return 3
        if ERROR_MISSING_INFO_IN_TC:
            return 1
        if ERROR_STATUS_TC_FOUND:
            return 2
        return 0

    @staticmethod
    def Detailed(reportDirectory, files):
        ws = 0
        Config = ConfigParser.ConfigParser()

        today = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
        ReportFileName = reportDirectory + r"\\" + today + "_Detail.xlsx"

        # Initialization of the report file in Excel
        # Open the workbook
        ReportWorkbook = Workbook()

        for report in files:
            ws += 1
            Config.read(report)

            # create a new sheet
            ReportWorksheet = ReportWorkbook.create_sheet("sheet" + str(ws))

            # Set the column width
            # ReportWorksheet.column_dimensions['A'].width = 30

            # set the sheet to landscape
            ReportWorksheet.page_setup.orientation = ReportWorksheet.ORIENTATION_LANDSCAPE

            # set the header and footer of the page
            ReportWorksheet.header_footer.center_footer.text = "&[Page]/&[Pages]"
            ReportWorksheet.header_footer.right_footer.text = "LIEBHERR MACHINE BULLE"
            ReportWorksheet.header_footer.left_footer.text = "&D\nDoc ID: &[File]"

            # Merge the cells
            # ReportWorksheet.merge_cells("A1:K1")

            row = 1
            column = 1
            ReportWorksheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=10)
            localCell = ReportWorksheet.cell(row=row, column=column)
            localCell.value = "GENERAL INFORMATION"
            localCell.font = Font(size=13, bold=True)
            localCell.alignment = Alignment(horizontal="center")
            localCell.fill = PatternFill(fill_type="solid", start_color="00EAEAEA")
            for c in range(1, 11):
                localCell = ReportWorksheet.cell(row=row, column=c)
                localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

            for subtitle in [("TestCase ID", "test case number"),
                             ("Requirement(s) cover", "requirement cover"),
                             ("Test Date", "test execution date"),
                             ("Test time", "test execution time"),
                             ("Test Case Description", "test case description"),
                             ("Result", "result"),
                             ("Tester note", "tester note"),
                             ("Incident No", "incident number")]:
                row += 1
                localCell = ReportWorksheet.cell(row=row, column=column)
                localCell.value = subtitle[0]
                localCell.font = Font(bold=True)
                localCell.style.alignment.wrap_text = True
                # localCell.fill = PatternFill(fill_type="solid", start_color="00EAEAEA")
                for c in range(1, 11):
                    localCell = ReportWorksheet.cell(row=row, column=c)
                    localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))

                column += 2
                localCell = ReportWorksheet.cell(row=row, column=column)
                localCell.value = Config.get(section="GENERIC", option=subtitle[1])
                column = 1

            row += 1
            ReportWorksheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=10)
            localCell = ReportWorksheet.cell(row=row, column=column)
            localCell.value = "ENVIRONMENT"
            localCell.font = Font(size=13, bold=True)
            localCell.alignment = Alignment(horizontal="center")
            localCell.fill = PatternFill(fill_type="solid", start_color="00EAEAEA")
            for c in range(1, 11):
                localCell = ReportWorksheet.cell(row=row, column=c)
                localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

            for subtitle in [("PC Name", "pc name"),
                             ("SW Number", "sw number"),
                             ("SW Revision", "sw revision"),
                             ("HIL model", "hil model"),
                             ("a2l master", "a2l master"),
                             ("a2l slave", "a2l slave"),
                             ("cdd", "cdd"),
                             ("CAN1 prot.", "can1 protocol"),
                             ("CAN2 prot.", "can2 protocol"),
                             ("Eng. variant", "engine variant"),
                             ("Ecu ID", "ecu id"),
                             ("Ecu spf_idx", "ecu spf_idx"),
                             ("Ecu Serial No", "ecu serial_number")]:
                row += 1
                localCell = ReportWorksheet.cell(row=row, column=column)
                localCell.value = subtitle[0]
                localCell.font = Font(bold=True)
                localCell.style.alignment.wrap_text = True
                # localCell.fill = PatternFill(fill_type="solid", start_color="00EAEAEA")
                for c in range(1, 11):
                    localCell = ReportWorksheet.cell(row=row, column=c)
                    localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))

                column += 2
                localCell = ReportWorksheet.cell(row=row, column=column)
                try:
                    localCell.value = Config.get(section="ENVIRONMENT", option=subtitle[1])
                except:
                    localCell.value = "NA"
                column = 1

            row += 1
            ReportWorksheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=10)
            localCell = ReportWorksheet.cell(row=row, column=column)
            localCell.value = "ACTUAL RESULT"
            localCell.font = Font(size=13, bold=True)
            localCell.alignment = Alignment(horizontal="center")
            localCell.fill = PatternFill(fill_type="solid", start_color="00EAEAEA")
            for c in range(1, 11):
                localCell = ReportWorksheet.cell(row=row, column=c)
                localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

            for line in Config.get(section="ACTUAL RESULTS", option="Log").split("\n"):
                row += 1
                localCell = ReportWorksheet.cell(row=row, column=column)
                localCell.value = line
                if line.find("NOT OK") > -1:
                    localCell.font = Font(color="00FF0000")

        # Remove the default worksheet
        ws = ReportWorkbook.get_sheet_by_name("Sheet")
        ReportWorkbook.remove_sheet(ws)
        # Save the report
        ReportWorkbook.save(ReportFileName)

    @staticmethod
    def TraceabilityMatrix(reportDirectory, files):
        """
        Generate an excel report in the folder reportDirectory using the files path given as argument
        :param reportDirectory: where the excel will be stored.
        :param files: list of path for the result files
        :return: None
        """

        # Open the excel file
        ReportWorkbook = Workbook()
        ReportWorksheet = ReportWorkbook.worksheets[0]

        # set the sheet name
        ReportWorksheet.title = "TraceMatrix"

        # set the header and footer of the page and page orientation
        # ReportWorksheet.header_footer.right_header.text = "&D"     # Add the date in the right of the header
        # ReportWorksheet.header_footer.left_header.text = "Traceability Matrix"
        ReportWorksheet.header_footer.right_footer.text = "LIEBHERR MACHINE BULLE"
        ReportWorksheet.header_footer.left_footer.text = "&D\nDoc ID: &[File]"
        ReportWorksheet.header_footer.center_footer.text = "&[Page]/&[Pages]"
        ReportWorksheet.page_setup.orientation = ReportWorksheet.ORIENTATION_LANDSCAPE

        # Set the first row to be reapeated at the top on each printed page
        ReportWorksheet.add_print_title(3)

        # Set the Main title
        localCell = ReportWorksheet.cell('A1')
        localCell.value = "Traceability Matrix"
        # Font properties
        localCell.font = Font(name='Calibri', size=16, bold=True)
        localCell.alignment = Alignment(horizontal="center")
        # Cell background
        localCell.fill = PatternFill(fill_type="solid", start_color="00F5F6F7")
        # Merge the cells
        ReportWorksheet.merge_cells("A1:E1")

        # Set the columns width
        ReportWorksheet.column_dimensions['A'].width = 50
        ReportWorksheet.column_dimensions['B'].width = 50
        ReportWorksheet.column_dimensions['C'].width = 30
        ReportWorksheet.column_dimensions['E'].width = 75

        # Set the headers of the table
        RowIndex = 3
        localCell = ReportWorksheet.cell(row=RowIndex, column=1)
        localCell.value = "Requirement"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=2)
        localCell.value = "TestCase/s"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=3)
        localCell.value = "Pass ratio"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=4)
        localCell.value = "Status"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=5)
        localCell.value = "Comment/s"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        # prepare the excel file name
        today = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
        ReportFileName = reportDirectory + r"\\" + today + "_TraceMatrix.xlsx"

        # initialize the parser
        Config = ConfigParser.ConfigParser()

        # create a dictionary starting from the data
        dictionary_of_requirement = {}
        dictionary_of_results = {}
        dictionary_of_comments = {}
        for report in files:
            Config.read(report)

            tc_id = Config.get(section="GENERIC", option="test case number")
            requirement = Config.get(section="GENERIC", option="requirement cover")
            result = Config.get(section="GENERIC", option="result")
            comment = Config.get(section="GENERIC", option="comment")
            if requirement != "":
                # split in lines
                requirement = requirement.split("\n")
                # remove first void line
                requirement = requirement[1:]
                for r in requirement:
                    if r not in dictionary_of_requirement.keys():
                        dictionary_of_requirement[str(r)] = []
                        dictionary_of_results[str(r)] = []
                        dictionary_of_comments[str(r)] = []
                    dictionary_of_requirement[str(r)].append(tc_id)
                    dictionary_of_results[str(r)].append(result)
                    if comment not in dictionary_of_comments[str(r)] and comment != "":  # if comment no already in dictionary and not empty -> add it
                        dictionary_of_comments[str(r)].append(comment)

        RowIndex = RowIndex + 1

        # Start to write on the excel the data contained in the dictionary
        for i in dictionary_of_requirement.keys():
            # Write the requirement
            localCell = ReportWorksheet.cell(row=RowIndex, column=1)
            localCell.value = i
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            localCell.alignment = Alignment(vertical="top")

            # write the Testcase/s
            localCell = ReportWorksheet.cell(row=RowIndex, column=2)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            value_to_write = ""
            for y in dictionary_of_requirement[i]:
                if value_to_write == "":
                    value_to_write = y
                else:
                    value_to_write = value_to_write + "\n" + y
            localCell.value = value_to_write

            # Write the pass ration
            localCell = ReportWorksheet.cell(row=RowIndex, column=3)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            total_test = len(dictionary_of_results[i])
            ok = 0
            nok = 0
            not_tested = 0
            for unit_result in dictionary_of_results[i]:
                if unit_result == "OK":
                    ok += 1
                elif unit_result == "NOT OK":
                    nok += 1
                elif unit_result == "NOT TESTED":
                    not_tested += 1
            ok = (ok*100)/total_test
            nok = (nok * 100) / total_test
            not_tested = (not_tested * 100) / total_test
            value_to_write = "OK: " + str(ok) + "%\n"
            value_to_write = value_to_write + "NOT OK: " + str(nok) + "%\n"
            value_to_write = value_to_write + "NOT TESTED: " + str(not_tested) + "%"
            localCell.value = value_to_write

            # Write the status
            localCell = ReportWorksheet.cell(row=RowIndex, column=4)
            localCell.alignment = Alignment(vertical="center", horizontal="center")
            if ok == 100:
                localCell.value = "OK"
                localCell.fill = PatternFill(fill_type="solid", start_color="0000FF00")
            elif nok > 0:
                localCell.value = "NOT OK"
                localCell.fill = PatternFill(fill_type="solid", start_color="00FF0000")
            elif not_tested == 100:
                localCell.value = "NOT TESTED"
            elif not_tested > 0 and nok == 0 and ok > 0:
                localCell.value = "PARTIAL"
                localCell.fill = PatternFill(fill_type="solid", start_color="00FFC000")

            # write the Comment/s
            localCell = ReportWorksheet.cell(row=RowIndex, column=5)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            value_to_write = ""
            for y in dictionary_of_comments[i]:
                if value_to_write == "":
                    value_to_write = "%s" %(y)
                else:
                    value_to_write = "%s \n%s" %(value_to_write,y)
            localCell.value = value_to_write

            RowIndex = RowIndex + 1
        # Save the report
        ReportWorkbook.save(ReportFileName)

    @staticmethod
    def IncidentMatrix(reportDirectory, files):
        """
        Generate an excel report in the folder reportDirectory using the files path given as argument
        :param reportDirectory: where the excel will be stored.
        :param files: list of path for the result files
        :return: None
        """

        # Open the excel file
        ReportWorkbook = Workbook()
        ReportWorksheet = ReportWorkbook.worksheets[0]

        # set the sheet name
        ReportWorksheet.title = "IncidentMatrix"

        # set the header and footer of the page and page orientation
        # ReportWorksheet.header_footer.right_header.text = "&D"     # Add the date in the right of the header
        # ReportWorksheet.header_footer.left_header.text = "Traceability Matrix"
        ReportWorksheet.header_footer.right_footer.text = "LIEBHERR MACHINE BULLE"
        ReportWorksheet.header_footer.left_footer.text = "&D\nDoc ID: &[File]"
        ReportWorksheet.header_footer.center_footer.text = "&[Page]/&[Pages]"
        ReportWorksheet.page_setup.orientation = ReportWorksheet.ORIENTATION_LANDSCAPE

        # Set the first row to be reapeated at the top on each printed page
        ReportWorksheet.add_print_title(3)

        # Set the Main title
        localCell = ReportWorksheet.cell('A1')
        localCell.value = "Incident Matrix"
        # Font properties
        localCell.font = Font(name='Calibri', size=16, bold=True)
        localCell.alignment = Alignment(horizontal="center")
        # Cell background
        localCell.fill = PatternFill(fill_type="solid", start_color="00F5F6F7")
        # Merge the cells
        ReportWorksheet.merge_cells("A1:F1")

        # Set the columns width
        ReportWorksheet.column_dimensions['A'].width = 50
        ReportWorksheet.column_dimensions['B'].width = 50
        ReportWorksheet.column_dimensions['C'].width = 50
        ReportWorksheet.column_dimensions['D'].width = 50
        ReportWorksheet.column_dimensions['E'].width = 50
        ReportWorksheet.column_dimensions['F'].width = 50

        # Set the headers of the table
        RowIndex = 3
        localCell = ReportWorksheet.cell(row=RowIndex, column=1)
        localCell.value = "Incident ID"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=2)
        localCell.value = "Incident title"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=3)
        localCell.value = "TestCase/s"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=4)
        localCell.value = "Requirement/s"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=5)
        localCell.value = "Assessment"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        localCell = ReportWorksheet.cell(row=RowIndex, column=6)
        localCell.value = "Blocking for release"
        localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        localCell.font = Font(name='Calibri', size=13, bold=True)

        # prepare the excel file name
        today = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
        ReportFileName = reportDirectory + r"\\" + today + "_IncidentMatrix.xlsx"

        # initialize the parser
        Config = ConfigParser.ConfigParser()

        # create a dictionary starting from the data
        dictionary_of_incident = {}

        user_entry_box = wx.TextEntryDialog(None, "Username", "Result viewer")
        ans = user_entry_box.ShowModal()
        if ans == wx.ID_OK:
            entered_user = user_entry_box.GetValue()
        else:
            entered_user = False
            return
        user_entry_box.Destroy()
        user = entered_user

        password_entry_box = wx.PasswordEntryDialog(None, "Password", "Result viewer")
        ans = password_entry_box.ShowModal()
        if ans == wx.ID_OK:
            entered_password = password_entry_box.GetValue()
        else:
            entered_password = False
            return
        password_entry_box.Destroy()
        password = entered_password

        for report in files:
            Config.read(report)

            tc_id = Config.get(section="GENERIC", option="test case number")
            requirement = Config.get(section="GENERIC", option="requirement cover")
            result = Config.get(section="GENERIC", option="result")
            incident = Config.get(section="GENERIC", option="incident number")

            # Fill the dictionary
            if incident != "":
                if incident not in dictionary_of_incident.keys():
                    dictionary_of_incident[str(incident)] = {}
                    dictionary_of_incident[str(incident)]["requirement"] = []
                    dictionary_of_incident[str(incident)]["testcase"] = []
                    # Get the title of the JIRA
                    cmd = "curl -v --user {0}:{1} https://jira.zdv.liebherr.i/browse/{2}".format(user,password,incident)
                    log = os.popen(cmd).readlines()
                    incident_title = ""
                    for l in log:
                        if l.find("summary-val") >= 0:
                            l = l.split('summary-val">')[1]
                            l = l.split("<")[0]
                            incident_title = l
                            break
                    dictionary_of_incident[str(incident)]["incident_title"] = incident_title
                dictionary_of_incident[str(incident)]["requirement"].append(requirement)
                dictionary_of_incident[str(incident)]["testcase"].append(tc_id)

        RowIndex = RowIndex + 1

        # Start to write on the excel the data contained in the dictionary
        for i in dictionary_of_incident.keys():
            # Write the incident
            localCell = ReportWorksheet.cell(row=RowIndex, column=1)
            localCell.value = i
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            localCell.alignment = Alignment(vertical="top")

            # write the Incident title
            localCell = ReportWorksheet.cell(row=RowIndex, column=2)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            localCell.value = dictionary_of_incident[str(i)]["incident_title"]

            # write the TestCase/s
            localCell = ReportWorksheet.cell(row=RowIndex, column=3)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            value_to_write = ""
            for y in dictionary_of_incident[i]["testcase"]:
                if value_to_write == "":
                    value_to_write = y
                else:
                    value_to_write = value_to_write + "\n" + y
            localCell.value = value_to_write

            # write the requirement/s
            localCell = ReportWorksheet.cell(row=RowIndex, column=4)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))
            value_to_write = ""
            for y in dictionary_of_incident[i]["requirement"]:
                if value_to_write == "":
                    value_to_write = y
                else:
                    value_to_write = value_to_write + "\n" + y
            localCell.value = value_to_write

            # write the assessment
            localCell = ReportWorksheet.cell(row=RowIndex, column=5)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))

            # write the Blocking for release
            localCell = ReportWorksheet.cell(row=RowIndex, column=5)
            localCell.alignment = Alignment(wrap_text=True, vertical="top")
            localCell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='hair'))

            RowIndex = RowIndex + 1
        # Save the report
        ReportWorkbook.save(ReportFileName)

class Application(wx.Frame):
    def __init__(self, parent, title):
        # Define useful variable
        self.dictionary_of_results = None
        self.html_creator = create_html.TC_result_HTML()
        self.logged_user = getpass.getuser()
        self.Config = ConfigParser.ConfigParser()
        self.three_quarter_of_screen_size = wx.DisplaySize()[0] * (3 / 4)

        self.setting = Setting(".data/setting.ini")

        self.result_list_column_idx_by_name = {}


        # Draw GUI
        wx.Frame.__init__(self, parent, title=title)  # , size=(200, 100))

        bitmap = wx.Bitmap('.data/SplashScreenLogo.png')
        splash = wx.adv.SplashScreen(bitmap, wx.adv.SPLASH_CENTER_ON_SCREEN | wx.adv.SPLASH_TIMEOUT, 1000, self)
        splash.Show()

        self.DrawGUI()

        # Logic stuff
        idx = 0
        categories = []
        for idx in range(self.result_list.GetColumnCount()):
            categories.append((self.result_list.GetColumn(idx)).GetText())
            idx = +1
        self.categories.AppendItems(categories)
        self.categories.SetLabel(categories[1])

        self.active_filter = {}
        for cat in categories:
            self.active_filter[cat] = []


        if self.logged_user.find("lmbhil") >= 0:
            self.active_filter["PC"].append(socket.gethostname())
        num_of_filter = 0
        for cat in self.active_filter.keys():
            num_of_filter = num_of_filter + len(self.active_filter[cat])
        self.active_filter_text.SetLabel("{0} filter/s".format(num_of_filter))
        self.active_filter_text.SetToolTip(str(self.active_filter))

        number_of_colum = self.result_list.GetColumnCount()
        for col_idx in range(number_of_colum):
            self.result_list_column_idx_by_name[self.result_list.GetColumn(col_idx).GetText()] = col_idx

        self.SetBackgroundColour('white')
        # self.Centre()  # center the window to the screen
        # self.Fit()  # fit the size of the window
        self.Maximize(True)
        self.Show(True)

    def DrawGUI(self):
        sizer = wx.GridBagSizer()

        self.DrawMenuBar()
        self.DrawCustomToolbar()
        self.DrawMessages()

        self.retrieve_data_from_folder = wx.DirPickerCtrl(self, message="Choose folder", style=wx.DIRP_USE_TEXTCTRL | wx.DIRP_SMALL, name="retrieve_data_from_folder")
        self.retrieve_data_from_folder.SetToolTip("Source folder for data")
        self.retrieve_data_from_folder.SetPath("Z:\Public\Transfert\HIL\Results\RawResults")

        self.result_store_folder = wx.DirPickerCtrl(self, message="Choose folder", style=wx.DIRP_USE_TEXTCTRL | wx.DIRP_SMALL)
        self.result_store_folder.SetInitialSize((400, -1))
        self.result_store_folder.SetToolTip("Choose the folder to store the results")

        self.result_list = wx.ListCtrl(self, style=wx.LC_REPORT | wx.BORDER_SUNKEN, name="result_list")
        self.result_list.InsertColumn(0, '#')
        self.result_list.SetColumnWidth(0, 30)
        self.result_list.InsertColumn(1, 'TestCase ID')
        self.result_list.SetColumnWidth(1, 300)
        self.result_list.InsertColumn(2, 'Description')
        self.result_list.SetColumnWidth(2, 300)
        self.result_list.InsertColumn(3, 'Result')
        self.result_list.InsertColumn(4, 'Comment')
        self.result_list.SetColumnWidth(4, 300)
        self.result_list.InsertColumn(5, 'Sw no')
        self.result_list.InsertColumn(6, 'Sw rev')
        self.result_list.InsertColumn(7, 'Incident No.')
        self.result_list.InsertColumn(8, 'Date')
        self.result_list.InsertColumn(9, 'Time')
        self.result_list.InsertColumn(10, 'PC')
        self.result_list.InsertColumn(11, 'Checksum cal/appl')
        self.result_list.InsertColumn(12, 'Path')
        self.result_list.SetColumnWidth(12, 0)
        self.result_list.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClicked)
        self.result_list.SetMaxSize(wx.Size(wx.DisplaySize()[0], -1))

        # pane = wx.CollapsiblePane(self, label="Metrics", style=wx.CP_DEFAULT_STYLE, name="Metrics")
        # pane.SetBackgroundColour("white")
        # pane.Expand()
        # pane.Bind(wx.EVT_COLLAPSIBLEPANE_CHANGED, self.OnPaneChanged)
        # self.metric_panel = pane.GetPane()
        self.DrawMetric()

        sizer.Add(self.ToolbarPanel, pos=(0, 0), flag=wx.ALL | wx.EXPAND, span=(1, 3))
        sizer.Add(wx.StaticText(self, label="Reading directory: "), pos=(1, 0), flag=wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_RIGHT, span=(1, 1))
        sizer.Add(self.retrieve_data_from_folder, pos=(1, 1), flag=wx.TOP | wx.EXPAND, border=3, span=(1, 1))
        sizer.Add(wx.StaticText(self, label="Saving directory: "), pos=(2, 0), flag=wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_RIGHT, span=(1, 1))
        sizer.Add(self.result_store_folder, pos=(2, 1), flag=wx.TOP | wx.EXPAND, border=3, span=(1, 1))
        sizer.Add(self.metric_panel, pos=(3, 0), flag=wx.ALL | wx.EXPAND, border=0, span=(1, 2))
        sizer.Add(self.result_list, pos=(4, 0), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.EXPAND, border=0, span=(1, 2))
        sizer.Add(self.MessagePanel, pos=(5, 0), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.EXPAND, border=0, span=(1, 2))

        sizer.AddGrowableCol(1)
        sizer.AddGrowableRow(4)
        self.SetSizer(sizer)
        self.Layout()

        # Create statusbar
        self.statusbar = self.CreateStatusBar(4)  # A Statusbar with 3 columns in the bottom of the window
        self.statusbar.SetStatusWidths([-1, -1, -6, -1])
        self.statusbar.SetStatusText("0 Item/s")
        # self.statusbar.SetStatusText("v. " + SVN.GetVersion(file_path=__file__), 3)

        self.ToggleVisibilityPanel()

    def DrawMenuBar(self):
        # create the menubar
        menubar = wx.MenuBar()

        # create menu/s
        fileMenu = wx.Menu()
        filtersMenu = wx.Menu()
        reportMenu = wx.Menu()
        windowsMenu = wx.Menu()
        helpMenu = wx.Menu()

        # populate menu/s
        read_data = wx.MenuItem(fileMenu, wx.ID_ANY, "Read...", "")
        img = wx.Image(r"Images\16px\binoculars.png", wx.BITMAP_TYPE_ANY)
        read_data.SetBitmap(wx.Bitmap(img))
        self.recursive = wx.MenuItem(fileMenu, wx.ID_ANY, "Recursive", kind=wx.ITEM_CHECK)
        reset_folder = wx.MenuItem(fileMenu, wx.ID_ANY, "Reset folder", "")
        img = wx.Image(r"Images\16px\folder2.png", wx.BITMAP_TYPE_ANY)
        reset_folder.SetBitmap(wx.Bitmap(img))
        save_data = wx.MenuItem(fileMenu, wx.ID_ANY, "Save...", "")
        img = wx.Image(r"Images\16px\Save2.png", wx.BITMAP_TYPE_ANY)
        save_data.SetBitmap(wx.Bitmap(img))
        reset_all_filters = wx.MenuItem(fileMenu, wx.ID_ANY, "Reset all", "Reset filters")
        self.case_sensitive_check = wx.MenuItem(filtersMenu, wx.ID_ANY, "Case sensitive", kind=wx.ITEM_CHECK)
        set_today = wx.MenuItem(filtersMenu, wx.ID_ANY, "Set today", "")
        set_this_pc = wx.MenuItem(filtersMenu, wx.ID_ANY, "Set this pc", "")
        set_hil1 = wx.MenuItem(filtersMenu, wx.ID_ANY, "Set HIL1", "")
        set_hil2 = wx.MenuItem(filtersMenu, wx.ID_ANY, "Set HIL2", "")
        set_hil3 = wx.MenuItem(filtersMenu, wx.ID_ANY, "Set HIL3", "")
        html_submenu = wx.Menu()
        excel_submenu = wx.Menu()
        html_report_info = wx.MenuItem(fileMenu, wx.ID_ANY, "Info", "")
        img = wx.Image(r"Images\16px\Create.png", wx.BITMAP_TYPE_ANY)
        html_report_info.SetBitmap(wx.Bitmap(img))
        html_report_other_restriction = wx.MenuItem(fileMenu, wx.ID_ANY, "Other restriction", "")
        img = wx.Image(r"Images\16px\Create.png", wx.BITMAP_TYPE_ANY)
        html_report_other_restriction.SetBitmap(wx.Bitmap(img))
        new_deviation = wx.MenuItem(fileMenu, wx.ID_ANY, "Deviation", "")
        img = wx.Image(r"Images\16px\Create.png", wx.BITMAP_TYPE_ANY)
        new_deviation.SetBitmap(wx.Bitmap(img))
        generate_report_html = wx.MenuItem(fileMenu, wx.ID_ANY, "Generate", "")
        img = wx.Image(r"Images\16px\technical-support.png", wx.BITMAP_TYPE_ANY)
        generate_report_html.SetBitmap(wx.Bitmap(img))
        generate_report = wx.MenuItem(fileMenu, wx.ID_ANY, "Generate", "")
        img = wx.Image(r"Images\16px\technical-support.png", wx.BITMAP_TYPE_ANY)
        generate_report.SetBitmap(wx.Bitmap(img))
        self.selected_reports = {}
        self.selected_reports["overview"] = wx.MenuItem(reportMenu, wx.ID_ANY, "Overview", kind=wx.ITEM_CHECK)
        self.selected_reports["trace"] = wx.MenuItem(reportMenu, wx.ID_ANY, "Trace Matrix", kind=wx.ITEM_CHECK)
        self.selected_reports["incident"] = wx.MenuItem(reportMenu, wx.ID_ANY, "Incident Matrix", kind=wx.ITEM_CHECK)
        self.selected_reports["details"] = wx.MenuItem(reportMenu, wx.ID_ANY, "Details", kind=wx.ITEM_CHECK)
        open_summary = wx.MenuItem(fileMenu, wx.ID_ANY, "Open Summary", "")
        img = wx.Image(r"Images\16px\Chart.png", wx.BITMAP_TYPE_ANY)
        open_summary.SetBitmap(wx.Bitmap(img))
        self.metric_panel_visibility = wx.MenuItem(windowsMenu, wx.ID_ANY, "Show metric", kind=wx.ITEM_CHECK)
        self.message_panel_visibility = wx.MenuItem(windowsMenu, wx.ID_ANY, "Show messages", kind=wx.ITEM_CHECK)
        manual = wx.MenuItem(helpMenu, wx.ID_ANY, "Manual", "")
        report_problem = wx.MenuItem(helpMenu, wx.ID_ANY, "Report problem", "")
        view_log_file = wx.MenuItem(helpMenu, wx.ID_ANY, "View log file", "")
        about = wx.MenuItem(helpMenu, wx.ID_ANY, "About", "")

        fileMenu.Append(read_data)
        fileMenu.Append(self.recursive)
        fileMenu.Append(reset_folder)
        fileMenu.AppendSeparator()
        fileMenu.Append(save_data)
        filtersMenu.Append(reset_all_filters)
        filtersMenu.Append(self.case_sensitive_check)
        filtersMenu.AppendSeparator()
        filtersMenu.Append(set_today)
        filtersMenu.Append(set_this_pc)
        filtersMenu.Append(set_hil1)
        filtersMenu.Append(set_hil2)
        filtersMenu.Append(set_hil3)

        reportMenu.Append(wx.ID_ANY, 'HTML', html_submenu)
        reportMenu.Append(wx.ID_ANY, 'Excel', excel_submenu)

        html_submenu.Append(html_report_info)
        html_submenu.Append(html_report_other_restriction)
        html_submenu.Append(new_deviation)
        html_submenu.AppendSeparator()
        html_submenu.Append(generate_report_html)
        excel_submenu.Append(open_summary)
        excel_submenu.Append(generate_report)
        excel_submenu.Append(self.selected_reports["overview"])
        excel_submenu.Append(self.selected_reports["trace"])
        excel_submenu.Append(self.selected_reports["incident"])
        excel_submenu.Append(self.selected_reports["details"])

        windowsMenu.Append(self.metric_panel_visibility)
        windowsMenu.Append(self.message_panel_visibility)
        helpMenu.Append(manual)
        helpMenu.Append(report_problem)
        helpMenu.Append(view_log_file)
        helpMenu.AppendSeparator()
        helpMenu.Append(about)

        self.recursive.Check()
        self.selected_reports["overview"].Check()
        self.selected_reports["trace"].Check()
        if self.setting.read(section="windows", option="metric panel") == "True":
            self.metric_panel_visibility.Check()
        if self.setting.read(section="windows", option="messages panel") == "True":
            self.message_panel_visibility.Check()

        reset_folder.Enable(False)

        # bind event to menu/s
        self.Bind(wx.EVT_MENU, lambda event: self.ReadDataandUpdatelist(), read_data)
        self.Bind(wx.EVT_MENU, lambda event: self.storeFile(), save_data)
        self.Bind(wx.EVT_MENU, lambda event: self.ResetAllFilters(), reset_all_filters)
        self.Bind(wx.EVT_MENU, lambda event: self.SetQuickFilter("today"), set_today)
        self.Bind(wx.EVT_MENU, lambda event: self.SetQuickFilter("this pc"), set_this_pc)
        self.Bind(wx.EVT_MENU, lambda event: self.SetQuickFilter("HIL1"), set_hil1)
        self.Bind(wx.EVT_MENU, lambda event: self.SetQuickFilter("HIL2"), set_hil2)
        self.Bind(wx.EVT_MENU, lambda event: self.SetQuickFilter("HIL3"), set_hil3)
        self.Bind(wx.EVT_MENU, lambda event: self.NewHTMLInfo(), html_report_info)
        self.Bind(wx.EVT_MENU, lambda event: self.NewOtherRestrictionFile(), html_report_other_restriction)
        self.Bind(wx.EVT_MENU, lambda event: self.NewDeviation(), new_deviation)
        self.Bind(wx.EVT_MENU, lambda event: self.GenerateReportHTML(), generate_report_html)
        self.Bind(wx.EVT_MENU, lambda event: self.GenerateReport(), generate_report)
        self.Bind(wx.EVT_MENU, lambda event: self.ViewReportTemplate(), open_summary)
        self.Bind(wx.EVT_MENU, lambda event: self.ToggleVisibilityPanel(), self.metric_panel_visibility)
        self.Bind(wx.EVT_MENU, lambda event: self.ToggleVisibilityPanel(), self.message_panel_visibility)
        self.Bind(wx.EVT_MENU, lambda event: self.Manual(), manual)
        self.Bind(wx.EVT_MENU, lambda event: self.ReportProblem(), report_problem)
        self.Bind(wx.EVT_MENU, lambda event: self.OpenLogFile(), view_log_file)
        self.Bind(wx.EVT_MENU, lambda event: self.AboutDialog(), about)

        # add menus to menubar
        menubar.Append(fileMenu, '&File')
        menubar.Append(filtersMenu, '&Filters')
        menubar.Append(reportMenu, '&Report')
        menubar.Append(windowsMenu, '&Window')
        menubar.Append(helpMenu, '&Help')

        self.SetMenuBar(menubar)

    def DrawCustomToolbar(self):
        self.ToolbarPanel = wx.Panel(self)  # Remove the border -> bd=0, relief='ridge'
        self.ToolbarPanel.SetBackgroundColour(wx.Colour(245, 246, 247))

        sizer = wx.GridBagSizer()

        bmp = wx.Bitmap(r"Images\32px\binoculars.png", wx.BITMAP_TYPE_PNG)
        Retreive = wx.BitmapButton(self.ToolbarPanel, id=wx.ID_ANY, bitmap=bmp, size=(bmp.GetWidth() + 3, bmp.GetHeight() + 3), name="retreive", style=wx.NO_BORDER)
        # Retreive = wx.Button(self.ToolbarPanel, label="Read", name="retreive")
        # Retreive.SetBitmap(wx.Bitmap(r"Images\32px\binoculars.png"), wx.LEFT)
        Retreive.SetToolTip("Read data from source")
        self.no_of_files = wx.SpinCtrl(self.ToolbarPanel, max=100000)
        self.no_of_files.SetInitialSize((60, -1))
        self.no_of_files.SetToolTip("Max numbers of files to read")
        self.no_of_files.SetValue(200)
        bmp = wx.Bitmap(r"Images\32px\folder.png", wx.BITMAP_TYPE_PNG)
        reset_folder = wx.BitmapButton(self.ToolbarPanel, id=wx.ID_ANY, bitmap=bmp, size=(bmp.GetWidth() + 3, bmp.GetHeight() + 3), name="Reset_folder", style=wx.NO_BORDER)
        searchByLbl = wx.StaticText(self.ToolbarPanel, label="Search By:")
        self.categories = wx.ComboBox(self.ToolbarPanel)
        self.search = wx.SearchCtrl(self.ToolbarPanel, style=wx.TE_PROCESS_ENTER)
        self.active_filter_text = wx.StaticText(self.ToolbarPanel, label="0 filter(s)")
        bmp = wx.Bitmap(r"Images\32px\save.png", wx.BITMAP_TYPE_PNG)
        store = wx.BitmapButton(self.ToolbarPanel, id=wx.ID_ANY, bitmap=bmp, size=(bmp.GetWidth() + 5, bmp.GetHeight() + 5), name="store", style=wx.NO_BORDER)
        store.SetToolTip("Save data to choosen location")
        bmp = wx.Bitmap(r"Images\32px\technical-support.png", wx.BITMAP_TYPE_PNG)
        generate_report = wx.BitmapButton(self.ToolbarPanel, id=wx.ID_ANY, bitmap=bmp, size=(bmp.GetWidth() + 5, bmp.GetHeight() + 5), name="generate_report", style=wx.NO_BORDER)
        generate_report.SetToolTip("Generate the selected report/s")
        bmp = wx.Bitmap(r"Images\32px\Chart.png", wx.BITMAP_TYPE_PNG)
        btn_template = wx.BitmapButton(self.ToolbarPanel, id=wx.ID_ANY, bitmap=bmp, size=(bmp.GetWidth() + 5, bmp.GetHeight() + 5), name="summary_template", style=wx.NO_BORDER)
        btn_template.SetToolTip("Open the summary page template")

        sizer.Add(Retreive, pos=(0, 0), flag=wx.LEFT | wx.ALIGN_CENTER_VERTICAL, border=5, span=(1, 1))
        sizer.Add(self.no_of_files, pos=(0, 1), flag=wx.LEFT | wx.ALIGN_CENTER_VERTICAL, border=10, span=(1, 1))
        sizer.Add(reset_folder, pos=(0, 2), flag=wx.LEFT | wx.ALIGN_CENTER_VERTICAL, border=10, span=(1, 1))
        sizer.Add(store, pos=(0, 3), flag=wx.LEFT, border=20, span=(1, 1))
        sizer.Add(wx.StaticLine(self.ToolbarPanel, style=wx.LI_VERTICAL), pos=(0, 4), flag=wx.RIGHT | wx.LEFT | wx.EXPAND, border=10, span=(1, 1))
        sizer.Add(searchByLbl, pos=(0, 5), flag=wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, border=1, span=(1, 1))
        sizer.Add(self.categories, pos=(0, 6), flag=wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, border=1, span=(1, 1))
        sizer.Add(self.search, pos=(0, 7), flag=wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, border=1, span=(1, 1))
        sizer.Add(self.active_filter_text, pos=(0, 8), flag=wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, border=1, span=(1, 1))
        sizer.Add(wx.StaticLine(self.ToolbarPanel, style=wx.LI_VERTICAL), pos=(0, 9), flag=wx.RIGHT | wx.LEFT | wx.EXPAND, border=10, span=(1, 1))
        sizer.Add(generate_report, pos=(0, 10), flag=wx.ALL, border=0, span=(1, 1))
        sizer.Add(btn_template, pos=(0, 11), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(1, 1))

        Retreive.Bind(wx.EVT_BUTTON, self.OnClicked)
        reset_folder.Bind(wx.EVT_BUTTON, self.OnClicked)
        store.Bind(wx.EVT_BUTTON, self.OnClicked)
        self.search.Bind(wx.EVT_TEXT_ENTER, self.OnSearch)
        generate_report.Bind(wx.EVT_BUTTON, lambda event: self.GenerateReportHTML())
        btn_template.Bind(wx.EVT_BUTTON, self.OnClicked)

        self.ToolbarPanel.SetSizer(sizer)

    def DrawMetric(self):
        self.metric_panel = wx.Panel(self)  # Remove the border -> bd=0, relief='ridge'
        self.metric_panel.SetBackgroundColour(wx.WHITE)

        sizer = wx.GridBagSizer()
        statistic_txt = wx.StaticText(self.metric_panel, wx.ID_ANY, label="Statistics")
        statistic_txt.SetBackgroundColour(wx.Colour(153, 180, 209))
        self.total_number_tc = wx.StaticText(self.metric_panel, wx.ID_ANY, label="0")
        self.total_number_tc.SetFont(wx.Font(12, wx.SWISS, wx.NORMAL, wx.BOLD))
        testcases_txt = wx.StaticText(self.metric_panel, wx.ID_ANY, label="TestCases")
        testcases_txt.SetFont(wx.Font(12, wx.SWISS, wx.NORMAL, wx.BOLD))
        self.tc_ok = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" OK")
        self.tc_ok.SetForegroundColour(wx.GREEN)
        self.tc_nok = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" NOT OK")
        self.tc_nok.SetForegroundColour(wx.RED)
        self.tc_not_tested = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" NOT TESTED")
        self.tc_not_tested.SetForegroundColour(wx.BLUE)
        self.tc_manual = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" MANUAL")
        self.tc_other = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" OTHER")
        self.defect = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" Defect(s)")
        self.testing_days = wx.StaticText(self.metric_panel, wx.ID_ANY, label=" Testing day(s)")


        sizer.Add(statistic_txt, pos=(0, 0), flag=wx.ALIGN_CENTER_HORIZONTAL | wx.EXPAND, border=0, span=(1, 7))
        sizer.Add(self.total_number_tc, pos=(1, 0), flag=wx.ALL | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(3, 1))
        sizer.Add(testcases_txt, pos=(4, 0), flag=wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, border=0, span=(1, 1))
        sizer.Add(wx.StaticLine(self.metric_panel, style=wx.LI_VERTICAL), pos=(1, 1), flag=wx.ALL | wx.EXPAND, border=10, span=(5, 1))
        sizer.Add(self.tc_ok, pos=(1, 2), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(1, 1))
        sizer.Add(self.tc_nok, pos=(2, 2), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(1, 1))
        sizer.Add(self.tc_not_tested, pos=(3, 2), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(1, 1))
        sizer.Add(self.tc_manual, pos=(4, 2), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(1, 1))
        sizer.Add(self.tc_other, pos=(5, 2), flag=wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=0, span=(1, 1))
        sizer.Add(wx.StaticLine(self.metric_panel, style=wx.LI_VERTICAL), pos=(1, 3), flag=wx.ALL | wx.EXPAND, border=10, span=(5, 1))
        sizer.Add(self.defect, pos=(1, 4), flag=wx.ALIGN_CENTER_VERTICAL, border=0, span=(5, 1))
        sizer.Add(wx.StaticLine(self.metric_panel, style=wx.LI_VERTICAL), pos=(1, 5), flag=wx.ALL | wx.EXPAND, border=10, span=(5, 1))
        sizer.Add(self.testing_days, pos=(1, 6), flag=wx.ALIGN_CENTER_VERTICAL, border=0, span=(5, 1))

        sizer.AddGrowableCol(6)

        self.metric_panel.SetSizer(sizer)

    def DrawMessages(self):
        self.MessagePanel = wx.Panel(self)  # Remove the border -> bd=0, relief='ridge'
        self.MessagePanel.SetBackgroundColour(wx.WHITE)

        sizer = wx.GridBagSizer()
        message_txt = wx.StaticText(self.MessagePanel, wx.ID_ANY, label="Messages")
        message_txt.SetBackgroundColour(wx.Colour(153, 180, 209))
        self.tree_ctrl = wx.TreeCtrl(self.MessagePanel, wx.ID_ANY)
        self.root_errors = self.tree_ctrl.AddRoot("Errors")
        # self.tree_ctrl.SetPyData(self.root, ("key", "value"))
        # self.tree_ctrl.AppendItem(self.root, '1st message')
        self.tree_ctrl.Expand(self.root_errors)


        sizer.Add(message_txt, pos=(0, 0), flag=wx.TOP | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_BOTTOM | wx.EXPAND, border=0, span=(1, 1))
        sizer.Add(self.tree_ctrl, pos=(1, 0), flag= wx.ALIGN_BOTTOM | wx.EXPAND, border=0, span=(1, 1))

        sizer.AddGrowableCol(0)
        self.MessagePanel.SetSizer(sizer)

    def ToggleVisibilityPanel(self):
        if self.message_panel_visibility.IsChecked():
            self.MessagePanel.Show()
            self.setting.write(section="windows", option="messages panel", value="True")
        elif not self.message_panel_visibility.IsChecked():
            self.MessagePanel.Hide()
            self.setting.write(section="windows", option="messages panel", value="False")

        if self.metric_panel_visibility.IsChecked():
            self.metric_panel.Show()
            self.setting.write(section="windows", option="metric panel", value="True")
        elif not self.metric_panel_visibility.IsChecked():
            self.metric_panel.Hide()
            self.setting.write(section="windows", option="metric panel", value="False")

        self.Layout()

    def OnClicked(self, event):
        btn = event.GetEventObject().GetName()
        # print "Label of pressed button = ", btn

        if btn == "retreive":
            self.ReadDataandUpdatelist()

        elif btn == "Reset_folder":
            self.retrieve_data_from_folder.SetPath("Z:\Public\Transfert\HIL\Results\RawResults")

        elif btn == "store":
            self.storeFile()

        elif btn == "generate_report":
            self.GenerateReport()

        elif btn == "summary_template":
            self.ViewReportTemplate()

    def GenerateReport(self):
        # Do not generate if source folder is the Z raw result
        if self.retrieve_data_from_folder.GetPath() == r"Z:\Public\Transfert\HIL\Results\RawResults":
            wx.MessageBox("Not possile to generate a report with this folder", "Wrong source folder", wx.ICON_INFORMATION)
            return

        # Get the showed item/s and create a list of the item path
        last_colum_number = self.result_list.GetColumnCount() - 1
        list_item = []
        item_index = -1
        while True:
            item_index = self.result_list.GetNextItem(item_index)
            if item_index == -1:
                break
            item = self.result_list.GetItem(item_index, last_colum_number)
            list_item.append(item.GetText())

        # Stop if the list is empty
        if not list_item:
            wx.MessageBox("No result are shown in the list", "Empty list", wx.ICON_INFORMATION)
            return

        # Generate the reports
        if self.selected_reports["overview"].IsChecked():
            overview_status = Report.Overview(self.retrieve_data_from_folder.GetPath(), list_item)
            if overview_status == 1:
                wx.MessageBox("Missing 'Comment' or 'JIRA' in testcase/s\n\nREPORT CANNOT BE USED FOR VALIDATION", "Report overview", wx.ICON_EXCLAMATION)
            elif overview_status == 2:
                wx.MessageBox("Wrong status in TestCase/s detected\n\nREPORT CANNOT BE USED FOR VALIDATION", "Report overview", wx.ICON_EXCLAMATION)
            elif overview_status == 3:
                wx.MessageBox("Missing 'Comment' or 'JIRA' in testcase/s\nWorng status in TestCase/s detected\n\nREPORT CANNOT BE USED FOR VALIDATION", "Report overview", wx.ICON_EXCLAMATION)
        if self.selected_reports["trace"].IsChecked():
            Report.TraceabilityMatrix(self.retrieve_data_from_folder.GetPath(), list_item)
        if self.selected_reports["incident"].IsChecked():
            Report.IncidentMatrix(self.retrieve_data_from_folder.GetPath(), list_item)
        if self.selected_reports["details"].IsChecked():
            Report.Detailed(self.retrieve_data_from_folder.GetPath(), list_item)

        os.startfile(self.retrieve_data_from_folder.GetPath())

    def NewHTMLInfo(self):
        dlg = wx.FileDialog(self, "Create Info in...",defaultFile="info",defaultDir=self.retrieve_data_from_folder.GetPath(), wildcard="CSV files (*.csv)|*.csv", style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        else:
            save_path = dlg.GetPath()
        dlg.Destroy()

        info = open(save_path, "w")
        info.write("Maturity;\n")
        info.write("Project;\n")
        info.close()

    def NewOtherRestrictionFile(self):
        dlg = wx.FileDialog(self, "Create Other restriction in...", defaultFile="other_restriction", defaultDir=self.retrieve_data_from_folder.GetPath(), wildcard="CSV files (*.csv)|*.csv", style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        else:
            save_path = dlg.GetPath()
        dlg.Destroy()

        other_restriction = open(save_path, "w")
        other_restriction.write("Restriction\n")
        other_restriction.write("FIRST RESTRICTION HERE\n")
        other_restriction.write("SECOND RESTRICTION HERE\n")
        other_restriction.write("THIRD RESTRICTION HERE\n")
        other_restriction.write("N RESTRICTION HERE\n")
        other_restriction.close()

    def NewDeviation(self):
        # Do not generate if source folder is the Z raw result
        if self.retrieve_data_from_folder.GetPath() == r"Z:\Public\Transfert\HIL\Results\RawResults":
            wx.MessageBox("Not possile to generate a deviation file with this folder", "Result Viewer", wx.ICON_INFORMATION)
            return

        defect_counter = []
        defect_dictionary = collections.OrderedDict()

        incident_no_column = self.result_list_column_idx_by_name["Incident No."]
        comment_column = self.result_list_column_idx_by_name["Comment"]

        item_index = -1
        while True:
            item_index = self.result_list.GetNextItem(item_index)
            if item_index == -1:
                break

            incident = self.result_list.GetItem(item_index, incident_no_column).GetText()
            comment = self.result_list.GetItem(item_index, comment_column).GetText()
            # count defect
            if incident != "":
                if incident not in defect_dictionary.keys():
                    defect_dictionary[incident] = comment
                else:
                    defect_dictionary[incident] = defect_dictionary[incident] + "\n" + comment
        print(defect_dictionary)

        dlg = wx.FileDialog(self, "Create Deviation in...",defaultFile="deviation_assesment", defaultDir=self.retrieve_data_from_folder.GetPath(),  wildcard="CSV files (*.csv)|*.csv", style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        else:
            save_path = dlg.GetPath()
        dlg.Destroy()


        deviation = open(save_path, "w")
        deviation.write('"JIRA (do not change)";"Blocking (yes/no)";Comment (CSM + AC);;TestCase comment (do not change, only for info)\n')
        for jira in defect_dictionary.keys():
            deviation.write('{0};yes;Write your comment here;;"{1}"\n'.format(jira, defect_dictionary[jira]))
        deviation.close()

    def GenerateReportHTML(self):
        # Do not generate if source folder is the Z raw result
        if self.retrieve_data_from_folder.GetPath() == r"Z:\Public\Transfert\HIL\Results\RawResults":
            wx.MessageBox("Not possile to generate a report with this folder", "Wrong source folder", wx.ICON_INFORMATION)
            return

        # if error present in errors message panel ask confirmation
        if self.tree_ctrl.GetCount() > 1:
            if wx.NO == wx.MessageBox("Errors present in messages panel, do you want do generate anyway?", "Result Viewer", wx.YES_NO | wx.ICON_INFORMATION):
                return

        # Get the showed item/s and create a list of the item path
        last_colum_number = self.result_list.GetColumnCount() - 1
        list_item = []
        item_index = -1
        while True:
            item_index = self.result_list.GetNextItem(item_index)
            if item_index == -1:
                break
            item = self.result_list.GetItem(item_index, last_colum_number)
            list_item.append(item.GetText())

        # Stop if the list is empty
        if not list_item:
            wx.MessageBox("No result are shown in the list", "Empty list", wx.ICON_INFORMATION)
            return


        progress_dlg = wx.ProgressDialog("Result Viewer", "Generating report...")
        progress_dlg.Pulse()
        try:
            ReportHTML.ReportHTML(self.retrieve_data_from_folder.GetPath(), list_item)
        except:
            wx.MessageBox("Error in function GenerateReportHTML()\n\n{0}".format(traceback.format_exc()), "Result viewer", wx.ICON_ERROR)
        progress_dlg.Destroy()

    def ViewReportTemplate(self):
        os.startfile(".data\AC_48_xxx_ECU_XX_XX_XX_TestSummaryReport.xltx")

    def OnSearch(self, event):
        if self.search.GetValue() not in self.active_filter[self.categories.GetValue()]:
            self.active_filter[self.categories.GetValue()].append(self.search.GetValue())
        num_of_filter = 0
        for cat in self.active_filter.keys():
            num_of_filter = num_of_filter + len(self.active_filter[cat])
        self.active_filter_text.SetLabel("{0} filter(s)".format(num_of_filter))
        self.active_filter_text.SetToolTip(str(self.active_filter))
        self.search.Clear()

        self.update_list()

    def ResetAllFilters(self):
        num_of_filter = 0
        for cat in self.active_filter.keys():
            self.active_filter[cat] = []  # reset filters for specific categorie
            num_of_filter = num_of_filter + len(self.active_filter[cat])
        self.active_filter_text.SetLabel("{0} filter/s".format(num_of_filter))

        self.update_list()

    def SetQuickFilter(self, arg):
        if arg == "this pc":
            self.active_filter["PC"].append(socket.gethostname())
        elif arg == "today":
            today = datetime.datetime.now()
            self.active_filter["Date"].append(today.strftime("%d/%m/%Y"))
        elif arg == "HIL1":
            self.active_filter["PC"].append("lmbpc0588")
        elif arg == "HIL2":
            self.active_filter["PC"].append("lmbpc0580")
        elif arg == "HIL3":
            self.active_filter["PC"].append("lmbpc0775")

        num_of_filter = 0
        for cat in self.active_filter.keys():
            num_of_filter = num_of_filter + len(self.active_filter[cat])
        self.active_filter_text.SetLabel("{0} filter/s".format(num_of_filter))
        self.active_filter_text.SetToolTip(str(self.active_filter))

        self.update_list()

    def OnPaneChanged(self, event):
        event.Skip()
        self.Layout()

    def OnRightClicked(self,event):
        btn = event.GetEventObject().GetName()
        # print "Name of pressed widget = ", btn

        if btn == "result_list":
            men = wx.Menu()

            opentxt = wx.MenuItem(men, wx.NewId(), 'Open txt')
            img = wx.Image(r'Images\16px\Text preview.png', wx.BITMAP_TYPE_ANY)
            opentxt.SetBitmap(wx.Bitmap(img))
            openhtml = wx.MenuItem(men, wx.NewId(), 'Open html')
            img = wx.Image(r'Images\16px\Text preview.png', wx.BITMAP_TYPE_ANY)
            openhtml.SetBitmap(wx.Bitmap(img))
            replace = wx.MenuItem(men, wx.NewId(), 'Replace')
            openfile = wx.MenuItem(men, wx.NewId(), 'Show in explorer')
            img = wx.Image(r'Images\16px\Folder.png', wx.BITMAP_TYPE_ANY)
            openfile.SetBitmap(wx.Bitmap(img))
            check_with_testcampaign = wx.MenuItem(men, wx.NewId(), 'Check with TestCampaign')
            save = wx.MenuItem(men, wx.NewId(), 'Save')
            img = wx.Image(r'Images\16px\Save2.png', wx.BITMAP_TYPE_ANY)
            save.SetBitmap(wx.Bitmap(img))
            delete = wx.MenuItem(men, wx.NewId(), 'Delete')
            img = wx.Image(r'Images\16px\Delete.png', wx.BITMAP_TYPE_ANY)
            delete.SetBitmap(wx.Bitmap(img))

            men.AppendItem(opentxt)
            men.AppendItem(openhtml)
            men.AppendSeparator()
            men.AppendItem(replace)
            men.AppendItem(check_with_testcampaign)
            men.AppendSeparator()
            men.AppendItem(openfile)
            men.AppendSeparator()
            men.AppendItem(save)
            men.AppendItem(delete)

            self.Bind(wx.EVT_MENU, self.openTxt, opentxt)
            self.Bind(wx.EVT_MENU, self.openHtml, openhtml)
            self.Bind(wx.EVT_MENU, lambda event: self.CheckWithTestCampaign(), check_with_testcampaign)
            self.Bind(wx.EVT_MENU, lambda event: self.Replace(), replace)
            self.Bind(wx.EVT_MENU, self.openFileLoaction, openfile)
            self.Bind(wx.EVT_MENU, self.deleteFile, delete)
            self.Bind(wx.EVT_MENU, lambda event: self.storeFile(), save)
            panel_pos = self.ScreenToClient(wx.GetMousePosition())
            self.PopupMenu(men, panel_pos)

    def openTxt(self,event):
        last_colum_number = self.result_list.GetColumnCount() - 1
        list_item = []
        # read the selected item/s and create a list of the item path
        item_index = -1
        while True:
            item_index = self.result_list.GetNextSelected(item_index)
            if item_index == -1:
                break
            item = self.result_list.GetItem(item_index, last_colum_number)
            list_item.append(item.GetText())

        # open the file/s
        for l in list_item:
            os.startfile(l)

        #i = self.result_list.GetFirstSelected()
        #item = self.result_list.GetItem(i,6)
        #print item.GetText()

    def openHtml(self,event):
        try:
            last_colum_number = self.result_list.GetColumnCount() - 1
            list_item = []
            # read the selected item/s and create a list of the item path
            item_index = -1
            while True:
                item_index = self.result_list.GetNextSelected(item_index)
                if item_index == -1:
                    break
                item = self.result_list.GetItem(item_index, last_colum_number)
                list_item.append(item.GetText())

            # open the file/s
            for l in list_item:
                file_to_open = self.html_creator.create_file(l)
                os.startfile(file_to_open)
        except:
            wx.MessageBox("Error in function openHTML()\n\n{0}".format(traceback.format_exc()), "Result viewer", wx.ICON_ERROR)

    def openFileLoaction(self,event):
        last_colum_number = self.result_list.GetColumnCount() - 1
        item_index = self.result_list.GetFirstSelected()
        item = self.result_list.GetItem(item_index, last_colum_number)
        os.system('explorer /select, "' + item.GetText() + '"')

    def deleteFile(self,event):
        last_colum_number = self.result_list.GetColumnCount() - 1
        if self.result_list.GetFirstSelected() == -1:
            wx.MessageBox("No TestCase selected", "Delete", wx.ICON_INFORMATION)
            return
        confirmation = wx.MessageBox("Are You Sure?", "Delete",wx.YES_NO | wx.ICON_EXCLAMATION)
        if confirmation == 2:  # Yes
            list_item = []
            # read the selected item/s and create a list of the item path
            item_index = -1
            while True:
                item_index = self.result_list.GetNextSelected(item_index)
                if item_index == -1:
                    break
                item = self.result_list.GetItem(item_index, last_colum_number)
                list_item.append(item.GetText())
            try:
                for item in list_item:
                    os.remove(item)
                self.ReadDataandUpdatelist()
                self.statusbar.SetStatusText("File/s deleted", 1)
            except WindowsError:
                wx.MessageBox("Access to file/s denied", "Delete", wx.ICON_EXCLAMATION)

    def storeFile(self):
        last_colum_number = self.result_list.GetColumnCount() - 1
        if self.result_list.GetFirstSelected() == -1:
            wx.MessageBox("No TestCase selected", "Result viewer", wx.ICON_EXCLAMATION)
            return

        # check if the folder exist
        if not os.path.isdir(self.result_store_folder.GetPath()):
            wx.MessageBox("No valid directory choosen", "Result viewer", wx.ICON_EXCLAMATION)
            return

        store_folder = self.result_store_folder.GetPath()

        list_item = []
        # read the selected item/s and create a list of the item path
        item_index = -1
        while True:
            item_index = self.result_list.GetNextSelected(item_index)
            if item_index == -1:
                break
            item = self.result_list.GetItem(item_index, last_colum_number)
            list_item.append(item.GetText())

        # copy each item in the destination folder
        for item in list_item:
            shutil.copy(item, store_folder)

        # copy also the .mdf file
        error_message = ""
        for index in self.dictionary_of_results:
            try:
                if self.dictionary_of_results[index]["GENERIC"]["path"] in list_item:
                    p = self.dictionary_of_results[index]["GENERIC"]["path"].split("\\")[:-1]
                    p = "\\".join(p)
                    for m in self.dictionary_of_results[index]["GENERIC"]["measurement file/s"].split("\n"):
                        if m != "-":
                            shutil.copy(p + "\\" + m, store_folder)
            except:
                print("ERROR: an error raised with {0}".format(self.dictionary_of_results[index]["GENERIC"]["path"]))
                print(traceback.format_exc())
                error_message = error_message + "\nError raised with {0}".format(self.dictionary_of_results[index]["GENERIC"]["path"])

        if error_message != "":
            wx.MessageBox(error_message, "Result viewer", wx.ICON_EXCLAMATION)


            self.statusbar.SetStatusText("File/s stored",1)

    def Replace(self):
        # Check if at leat one item is selected
        if self.result_list.GetFirstSelected() == -1:
            wx.MessageBox("No TC selected", "Result viewer", wx.ICON_EXCLAMATION)
            return

        # use the first item to create list of key and option
        last_colum_number = self.result_list.GetColumnCount() - 1
        item_index = self.result_list.GetFirstSelected()
        item = self.result_list.GetItem(item_index, last_colum_number)

        self.Config = ConfigParser.ConfigParser()
        self.Config.read(item.GetText())
        sections = self.Config.sections()

        # ask to user the key
        entry_box = wx.SingleChoiceDialog(None, "Key", "Result viewer", choices=sections)
        ans = entry_box.ShowModal()
        if ans == wx.ID_OK:
            entered_key = entry_box.GetStringSelection()
        else:
            entered_key = False
            return
        entry_box.Destroy()

        # ask the user the option
        options = self.Config.options(entered_key)
        entry_box = wx.SingleChoiceDialog(None, "Item", "Result viewer", choices=options)
        ans = entry_box.ShowModal()
        if ans == wx.ID_OK:
            entered_item = entry_box.GetStringSelection()
        else:
            entered_item = False
            return
        entry_box.Destroy()

        # ask the user the new value to write
        entry_box = wx.TextEntryDialog(None, "New value", "Result viewer")
        ans = entry_box.ShowModal()
        if ans == wx.ID_OK:
            entered_new_value = entry_box.GetValue()
        else:
            entered_new_value = False
            return
        entry_box.Destroy()

        # read the selected item/s and create a list of the item path
        last_colum_number = self.result_list.GetColumnCount() - 1
        list_item = []
        item_index = -1
        while True:
            item_index = self.result_list.GetNextSelected(item_index)
            if item_index == -1:
                break
            item = self.result_list.GetItem(item_index, last_colum_number)
            list_item.append(item.GetText())

        for f in list_item:
            self.Config = ConfigParser.ConfigParser()
            self.Config.read(f)
            self.Config.set(section=entered_key, option=entered_item, value=entered_new_value)

            file_to_write = open(f, "w")
            self.Config.write(file_to_write)
            file_to_write.close()

        self.ReadDataandUpdatelist()

    def CheckWithTestCampaign(self):
        dlg = wx.FileDialog(self, "Choose test campaign...", wildcard="Text files (*.txt)|*.txt")
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        else:
            test_campaign_path = dlg.GetPath()
        dlg.Destroy()

        # open the test campaign and create a variable list with the tc
        f = open(test_campaign_path, "r")
        lines = f.readlines()
        f.close()

        test_campaign_list = []
        for line in lines:
            line = line.rstrip()  # remove the new line at the end
            test_campaign_list.append(line)

        # read the selected item/s and create a list TestCase ID
        list_item = []
        item_index = -1
        while True:
            item_index = self.result_list.GetNextItem(item_index)
            if item_index == -1:
                break
            item = self.result_list.GetItem(item_index, self.result_list_column_idx_by_name["TestCase ID"])
            list_item.append(item.GetText())

        missing_tc = []

        for test_camp in test_campaign_list:
            found = False
            for tc in list_item:
                if tc.find(test_camp) >= 0:
                    found = True
                    break

            if not found:
                missing_tc.append(test_camp)

        wx.MessageBox(str(missing_tc), "Result viewer", wx.ICON_NONE)

    def onMenuClick(self, event):
        btn = event.GetId()
        print("Menu clicker btn: ", btn)

    def ReadDataandUpdatelist(self):
        progress_dlg = wx.ProgressDialog("Read data", "Reading data from folder.")
        progress_dlg.Pulse()
        self.dictionary_of_results = self.get_result_files(self.retrieve_data_from_folder.GetPath(), self.no_of_files.GetValue())
        # print self.retrieve_data_from_folder.GetTextCtrlValue()
        self.update_list()
        progress_dlg.Destroy()

    def get_result_files(self, path="", number_of_file_to_read=10000):
        """
        Return a dictionary with all info of test case in a folder given as argument
        :param number_of_file_to_read:
        :return:
        """
        if path == "":
            print("No path available")
            return
        if not os.path.exists(path):
            wx.MessageBox("Path given not valid", "Result viewer", wx.ICON_EXCLAMATION)
            return

        a = TestResultParser(path=path, limit_number_of_file=int(self.no_of_files.GetValue()), recursive=self.recursive.IsChecked())
        dictionary, list_of_errors = a.get_dictionary()
        if list_of_errors != []:
            for error in list_of_errors:
                # logging.error("Following file contain error: {0}".format(error))
                print("Following file contain error: {0}".format(error))
        return dictionary

    def update_list(self):
        try:
            number_of_tc_on_screen = 0

            if self.dictionary_of_results == None:
                print("No RawResult to analyze")
                return

            self.result_list.DeleteAllItems()

            lst = []

            # Parse the created dictionary
            idx = 0
            for index in self.dictionary_of_results.keys():
                write_this_item = True

                try:
                    if self.case_sensitive_check.IsChecked():
                        PAR = re.DOTALL
                    else:
                        PAR = re.IGNORECASE | re.DOTALL

                    if write_this_item:
                        if self.active_filter["TestCase ID"] != []:
                            write_this_item = False
                            for filter in self.active_filter["TestCase ID"]:
                                if re.search(filter, self.dictionary_of_results[index]["GENERIC"]["test case number"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Description"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Description"]:
                                if re.search(filter, self.dictionary_of_results[index]["GENERIC"]["test case description"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Result"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Result"]:
                                if re.match(filter, self.dictionary_of_results[index]["GENERIC"]["result"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Comment"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Comment"]:
                                if re.search(filter, self.dictionary_of_results[index]["GENERIC"]["comment"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Sw no"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Sw no"]:
                                if re.search(filter, self.dictionary_of_results[index]["ENVIRONMENT"]["sw number"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Sw rev"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Sw rev"]:
                                if re.search(filter, self.dictionary_of_results[index]["ENVIRONMENT"]["sw revision"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Incident No."] != []:
                            write_this_item = False
                            for filter in self.active_filter["Incident No."]:
                                if re.search(filter, self.dictionary_of_results[index]["GENERIC"]["incident number"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Date"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Date"]:
                                if re.search(filter, self.dictionary_of_results[index]["GENERIC"]["test execution date"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["PC"] != []:
                            write_this_item = False
                            for filter in self.active_filter["PC"]:
                                if re.search(filter, self.dictionary_of_results[index]["ENVIRONMENT"]["pc name"], PAR) != None:
                                    write_this_item = True
                                    break

                    if write_this_item:
                        if self.active_filter["Checksum cal/appl"] != []:
                            write_this_item = False
                            for filter in self.active_filter["Checksum cal/appl"]:
                                if re.search(filter, self.dictionary_of_results[index]["ENVIRONMENT"]["checksum calibration and application"], PAR) != None:
                                    write_this_item = True
                                    break


                    if not write_this_item:
                        continue

                    if write_this_item:
                        if self.dictionary_of_results[index]["GENERIC"]["result"] == "OK":
                            r = "OK"
                        elif self.dictionary_of_results[index]["GENERIC"]["result"] == "NOT OK":
                            r = "NOK"
                        elif self.dictionary_of_results[index]["GENERIC"]["result"] == "MANUAL":
                            r = "MANUAL"
                        else:
                            r = "OTHER"

                        self.result_list.InsertItem(idx, str(idx))
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["TestCase ID"], self.dictionary_of_results[index]["GENERIC"]["test case number"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Description"], self.dictionary_of_results[index]["GENERIC"]["test case description"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Result"], self.dictionary_of_results[index]["GENERIC"]["result"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Comment"], self.dictionary_of_results[index]["GENERIC"]["comment"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Sw no"], self.dictionary_of_results[index]["ENVIRONMENT"]["sw number"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Sw rev"], self.dictionary_of_results[index]["ENVIRONMENT"]["sw revision"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Incident No."], self.dictionary_of_results[index]["GENERIC"]["incident number"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Date"], self.dictionary_of_results[index]["GENERIC"]["test execution date"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Time"], self.dictionary_of_results[index]["GENERIC"]["test execution time"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["PC"], self.dictionary_of_results[index]["ENVIRONMENT"]["pc name"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Checksum cal/appl"], self.dictionary_of_results[index]["ENVIRONMENT"]["checksum calibration and application"])
                        self.result_list.SetItem(idx, self.result_list_column_idx_by_name["Path"], self.dictionary_of_results[index]["GENERIC"]["path"])
                        if r == "OK":
                            self.result_list.SetItemBackgroundColour(idx, "green")
                        elif r == "NOK":
                            self.result_list.SetItemBackgroundColour(idx, "red")
                        elif r == "MANUAL":
                            self.result_list.SetItemBackgroundColour(idx, "yellow")

                        idx = idx + 1
                        number_of_tc_on_screen += 1
                except:
                    print(self.dictionary_of_results[index]["GENERIC"]["path"],)
                    print("generated a problem")
                    print(traceback.format_exc())


            self.StatusBar.SetStatusText(str(number_of_tc_on_screen)  + " Item/s")

            self.Updatemetrics()
        except:
            wx.MessageBox("Error in function update_list()\n\n{0}".format(traceback.format_exc()), "Result viewer", wx.ICON_ERROR)

    def Updatemetrics(self):
        """
        read item/s and create metrics
        """
        ready_for_generation = True
        missing_incident_number_for_nok = 0
        missing_comment_for_nok = 0
        missing_comment_for_not_tested = 0
        total_tc = 0
        ok_tc = 0
        nok_tc = 0
        not_tested_tc = 0
        manual_tc = 0
        other= 0
        defect_counter = []
        dates = []
        result_column       = self.result_list_column_idx_by_name["Result"]
        incident_no_column  = self.result_list_column_idx_by_name["Incident No."]
        comment_column      = self.result_list_column_idx_by_name["Comment"]
        date_column         = self.result_list_column_idx_by_name["Date"]
        checksum_column     = self.result_list_column_idx_by_name["Checksum cal/appl"]
        checksum_list = []
        self.tree_ctrl.DeleteChildren(self.root_errors)

        item_index = -1
        while True:
            item_index = self.result_list.GetNextItem(item_index)
            if item_index == -1:
                break

            tmp_result      = self.result_list.GetItem(item_index, result_column).GetText()
            tmp_incident_no = self.result_list.GetItem(item_index, incident_no_column).GetText()
            tmp_comment     = self.result_list.GetItem(item_index, comment_column).GetText()
            tmp_date        = self.result_list.GetItem(item_index, date_column).GetText()
            tmp_checksum    = self.result_list.GetItem(item_index, checksum_column).GetText()

            # count result
            total_tc += 1
            if tmp_result == "OK":
                ok_tc +=1
            elif tmp_result == "NOT OK":
                nok_tc += 1
            elif tmp_result == "NOT TESTED":
                not_tested_tc +=1
            elif tmp_result == "MANUAL":
                manual_tc += 1
                ready_for_generation = False
                self.tree_ctrl.AppendItem(self.root_errors, 'MANUAL status detected - row {0}'.format(item_index))
            else:
                other += 1
                ready_for_generation = False
                self.tree_ctrl.AppendItem(self.root_errors, 'Unknown status detected - row {0}'.format(item_index))

            # check incident number
            if tmp_result == "NOT OK":
                if tmp_incident_no == "":
                    missing_incident_number_for_nok += 1
                    ready_for_generation = False
                    self.tree_ctrl.AppendItem(self.root_errors, 'Missing incident number for NOT OK detected - row {0}'.format(item_index))

            # check comment
            if tmp_result == "NOT OK":
                if tmp_comment == "":
                    missing_comment_for_nok += 1
                    ready_for_generation = False
                    self.tree_ctrl.AppendItem(self.root_errors, 'Missing comment for NOT OK detected - row {0}'.format(item_index))

            if tmp_result == "NOT TESTED":
                if tmp_comment == "":
                    missing_comment_for_not_tested += 1
                    ready_for_generation = False
                    self.tree_ctrl.AppendItem(self.root_errors, 'Missing comment for NOT TESTED detected - row {0}'.format(item_index))

            # count defect
            if tmp_incident_no != "":
                if tmp_incident_no not in defect_counter:
                    defect_counter.append(tmp_incident_no)

            # add date
            dates.append(datetime.datetime.strptime(tmp_date,'%d/%m/%Y'))

            # add checksum
            checksum_list.append(tmp_checksum)


        if total_tc != 0:
            percent_ok = (100.0/total_tc)*ok_tc
            percent_nok = (100.0/total_tc)*nok_tc
            percent_not_tested = (100.0/total_tc)*not_tested_tc
            percent_manual = (100.0/total_tc)*manual_tc
            percent_other = (100.0/total_tc)*other
        else:
            percent_ok = 0
            percent_nok = 0
            percent_not_tested = 0
            percent_manual = 0
            percent_other = 0

        self.total_number_tc.SetLabel("{0}".format(total_tc))
        self.tc_ok.SetLabel("{0} | {1:.2f}% | OK".format(ok_tc, percent_ok))
        self.tc_nok.SetLabel("{0} | {1:.2f}% | NOT OK".format(nok_tc, percent_nok))
        self.tc_not_tested.SetLabel("{0} | {1:.2f}% | NOT TESTED".format(not_tested_tc, percent_not_tested))
        self.tc_manual.SetLabel("{0} | {1:.2f}% | MANUAL".format(manual_tc, percent_manual))
        self.tc_other.SetLabel("{0} | {1:.2f}% | OTHER".format(other, percent_other))

        self.defect.SetLabel("{0} Defect(s)".format(len(defect_counter)))

        if total_tc != 0:
            oldest_date     = max(dates)
            youngest_date   = min(dates)
            delta_date      = oldest_date - youngest_date
            delta_date      = delta_date.days
            delta_date      += 1
        else:
            delta_date = 0
        self.testing_days.SetLabel("{0} Testing day(s)".format(delta_date))

        # check all checksum are the same
        if len(set(checksum_list)) > 1:
            self.tree_ctrl.AppendItem(self.root_errors, "Different checksum: {0}".format(list(set(checksum_list))))

        if not ready_for_generation:
            self.tree_ctrl.Expand(self.root_errors)


        self.Layout()

    def Manual(self):
        os.startfile(r"C:\CAD\Workspace\HILSimulator\HIL_Tools\TestCase_RawResult_Manager\.data\Manual\Manual.pdf")

    def ReportProblem(self):
        os.startfile("https://github.com/Lexcere/ResultViewer/issues/new")

    def OpenLogFile(self):
        try:
            os.startfile(r".data\Log\{0}".format(log_file_name))
        except:
            wx.MessageBox(traceback.format_exc(), "Result Viewer", wx.ICON_ERROR)

    def AboutDialog(self):
        about = wx.adv.AboutDialogInfo()
        about.Name = "Result Viewer"
        about.Description = "Used to view/open/manange TestCase result"
        # about.Version = SVN.GetVersion(file_path=__file__)
        about.SetWebSite("www.pornhub.com")
        about.Developers = ["Cere"]
        about.AddArtist("Gabbo")
        wx.adv.AboutBox(about)

app = wx.App(False)
frame = Application(None, "Result Viewer")
app.MainLoop()

